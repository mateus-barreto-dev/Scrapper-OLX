import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def criar_planilha_modelo(filename="modelo_testes_olx.xlsx"):
    # Dados fictícios para teste de estrutura
    dados = [
        {
            "Ganho Potencial (R$)": 650.00,
            "Preço Anunciado (R$)": 1150.00,
            "Preço Estimado Revenda (R$)": 1800.00,
            "Margem (%)": 0.565,
            "Modelo / Processador": "Lenovo IdeaPad i5 10ª 8GB SSD 256GB",
            "Score": 9,
            "Condição & Alertas": "Excelente estado, vendedor com pressa",
            "Ação Sugerida": "Comprar Imediatamente",
            "Link": "https://www.olx.com.br",
            "Data": "2026-07-31"
        },
        {
            "Ganho Potencial (R$)": 400.00,
            "Preço Anunciado (R$)": 1800.00,
            "Preço Estimado Revenda (R$)": 2200.00,
            "Margem (%)": 0.222,
            "Modelo / Processador": "Acer Nitro 5 GTX 1650 i5 10ª",
            "Score": 8,
            "Condição & Alertas": "Marcas leves de uso na tampa",
            "Ação Sugerida": "Negociar Preço",
            "Link": "https://www.olx.com.br",
            "Data": "2026-07-31"
        },
        {
            "Ganho Potencial (R$)": -200.00,
            "Preço Anunciado (R$)": 2700.00,
            "Preço Estimado Revenda (R$)": 2500.00,
            "Margem (%)": -0.074,
            "Modelo / Processador": "Samsung Book i5 11ª 8GB",
            "Score": 3,
            "Condição & Alertas": "Anúncio superestimado acima do mercado",
            "Ação Sugerida": "Descartar",
            "Link": "https://www.olx.com.br",
            "Data": "2026-07-31"
        }
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Oportunidades"

    # Cabeçalhos
    headers = list(dados[0].keys())
    ws.append(headers)

    # Estilo do Cabeçalho
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Inserção de Dados e Formatação
    for row_num, item in enumerate(dados, 2):
        ws.cell(row=row_num, column=1, value=item["Ganho Potencial (R$)"]).number_format = 'R$ #,##0.00'
        ws.cell(row=row_num, column=2, value=item["Preço Anunciado (R$)"]).number_format = 'R$ #,##0.00'
        ws.cell(row=row_num, column=3, value=item["Preço Estimado Revenda (R$)"]).number_format = 'R$ #,##0.00'
        ws.cell(row=row_num, column=4, value=item["Margem (%)"]).number_format = '0.0%'
        ws.cell(row=row_num, column=5, value=item["Modelo / Processador"])
        ws.cell(row=row_num, column=6, value=item["Score"]).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=7, value=item["Condição & Alertas"])
        ws.cell(row=row_num, column=8, value=item["Ação Sugerida"])
        
        # Hyperlink no Excel
        cell_link = ws.cell(row=row_num, column=9, value="Acessar OLX")
        cell_link.hyperlink = item["Link"]
        cell_link.font = Font(color="0000FF", underline="single")
        
        ws.cell(row=row_num, column=10, value=item["Data"]).alignment = Alignment(horizontal="center")

        # Colorir célula do Ganho Potencial
        ganho = item["Ganho Potencial (R$)"]
        cell_ganho = ws.cell(row=row_num, column=1)
        if ganho >= 300:
            cell_ganho.fill = PatternFill(start_color="C6EFCE", fill_type="solid") # Verde
            cell_ganho.font = Font(color="006100", bold=True)
        elif ganho < 0:
            cell_ganho.fill = PatternFill(start_color="FFC7CE", fill_type="solid") # Vermelho
            cell_ganho.font = Font(color="9C0006")

    # Ajuste automático de largura de colunas
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(filename)
    print(f"Planilha '{filename}' gerada com sucesso!")

if __name__ == "__main__":
    criar_planilha_modelo()